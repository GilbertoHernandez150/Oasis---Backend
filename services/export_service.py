"""
export_service.py
Genera archivos Excel (.xlsx) con los datos del análisis.
Ubicación: backend/services/export_service.py
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Paleta OASIS ──────────────────────────────────────────
C_DARK_BG   = "0A0F1E"   # fondo principal
C_SURFACE   = "162032"   # superficie de tarjetas
C_TEAL      = "00D4AA"   # acento principal
C_BLUE      = "3B82F6"   # acento secundario
C_GREEN     = "10B981"   # positivo
C_RED       = "EF4444"   # negativo / gastos
C_AMBER     = "F59E0B"   # advertencia / impuestos
C_PURPLE    = "8B5CF6"   # indicadores
C_WHITE     = "FFFFFF"
C_GRAY_LIGHT= "E2E8F0"
C_GRAY_MED  = "64748B"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_WHITE, size=11, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _border():
    s = Side(style="thin", color="334155")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

def _header_row(ws, row, cols_data, bg_color=C_SURFACE, font_color=C_TEAL):
    """Escribe una fila de encabezado con estilo."""
    for col_idx, (text, width) in enumerate(cols_data, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill      = _fill(bg_color)
        cell.font      = _font(bold=True, color=font_color, size=10)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _data_row(ws, row, values, font_color=C_GRAY_LIGHT, bg_color=C_DARK_BG, bold=False):
    """Escribe una fila de datos con estilo."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill      = _fill(bg_color)
        cell.font      = _font(bold=bold, color=font_color, size=10)
        cell.alignment = _left() if col_idx == 1 else _right()
        cell.border    = _border()

def _title_row(ws, row, text, ncols, bg=C_DARK_BG, fg=C_TEAL, size=13):
    """Escribe un título que abarca varias columnas."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, color=fg, size=size)
    cell.alignment = _center()
    cell.border    = _border()
    ws.row_dimensions[row].height = 28

def _fmt_rd(value):
    """Formatea un valor como pesos dominicanos."""
    if value is None:
        return "—"
    try:
        return f"RD$ {float(value):,.2f}"
    except (TypeError, ValueError):
        return str(value)

def _set_tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color


# ══════════════════════════════════════════════════════════
# EXPORTAR ANÁLISIS ECONÓMICO
# ══════════════════════════════════════════════════════════
def export_economico_xlsx(data: dict, nombre_archivo: str = "análisis") -> bytes:
    """
    Genera un .xlsx con los resultados del análisis económico.
    Retorna los bytes del archivo para enviarlo como respuesta HTTP.
    """
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    _set_tab_color(ws1, C_TEAL)
    ws1.sheet_view.showGridLines = False

    # Encabezado OASIS
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value     = "🌴  OASIS — Análisis Financiero"
    c.fill      = _fill(C_DARK_BG)
    c.font      = _font(bold=True, color=C_TEAL, size=16)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:D2")
    c2 = ws1["A2"]
    c2.value     = f"Archivo: {nombre_archivo}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c2.fill      = _fill(C_SURFACE)
    c2.font      = _font(color=C_GRAY_MED, size=10, italic=True)
    c2.alignment = _center()
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 8

    # KPIs principales
    _title_row(ws1, 4, "📊  INDICADORES PRINCIPALES", 4)

    kpis = [
        ("INDICADOR",         "VALOR",                        "MONEDA", "DETALLE"),
        ("Total Ingresos",    data.get("total_ingresos", 0),  "RD$",    "Suma de todos los ingresos"),
        ("Total Gastos",      data.get("total_gastos", 0),    "RD$",    "Suma de todos los egresos"),
        ("Balance",           data.get("balance", 0),         "RD$",    "Ingresos - Gastos"),
        ("Total Transacciones", data.get("total_transacciones", 0), "—", "Cantidad de registros"),
        ("Promedio Ingreso",  data.get("promedio_ingreso", 0),"RD$",    "Promedio por transacción"),
        ("Promedio Gasto",    data.get("promedio_gasto", 0),  "RD$",    "Promedio por transacción"),
    ]

    _header_row(ws1, 5, [
        ("INDICADOR", 28), ("VALOR", 18), ("MONEDA", 10), ("DETALLE", 36)
    ])

    for i, (ind, val, mon, det) in enumerate(kpis[1:], 6):
        color_val = C_GREEN if ind == "Total Ingresos" else \
                    C_RED   if ind == "Total Gastos"   else \
                    (C_TEAL if (data.get("balance", 0) or 0) >= 0 else C_RED) \
                    if ind == "Balance" else C_GRAY_LIGHT
        _data_row(ws1, i, [ind, val, mon, det],
                  bg_color="0F172A" if i % 2 == 0 else C_DARK_BG)
        ws1.cell(i, 2).font      = _font(bold=True, color=color_val, size=11)
        ws1.cell(i, 2).number_format = '#,##0.00'

    # Impuestos
    imp = data.get("impuestos", {})
    tss = imp.get("tss", {})
    row = len(kpis) + 7

    ws1.row_dimensions[row].height = 8
    _title_row(ws1, row + 1, "🏦  IMPUESTOS REPÚBLICA DOMINICANA (Ley 87-01)", 4, fg=C_AMBER)

    _header_row(ws1, row + 2, [
        ("CONCEPTO", 28), ("MONTO (RD$)", 18), ("BASE", 10), ("NORMATIVA", 36)
    ], font_color=C_AMBER)

    imp_rows = [
        ("ISR (Impuesto Sobre la Renta)", imp.get("isr", 0),            "Ingresos",  "DGII — Tabla progresiva RD"),
        ("AFP (Fondo de Pensiones)",      tss.get("afp", 0),            "2.87%",     "Ley 87-01 Art. 186"),
        ("SFS (Seguro Familiar de Salud)",tss.get("sfs", 0),            "3.04%",     "Ley 87-01 Art. 118"),
        ("SRL (Riesgos Laborales)",       tss.get("srl", 0),            "0.10%",     "Ley 87-01 Art. 196"),
        ("TOTAL IMPUESTOS",               imp.get("total_impuestos", 0),"—",         "ISR + AFP + SFS + SRL"),
        ("INGRESO NETO",                  imp.get("ingreso_neto", 0),   "—",         "Ingresos - Total Impuestos"),
    ]

    for j, (concepto, monto, base, norma) in enumerate(imp_rows, row + 3):
        is_total = "TOTAL" in concepto or "NETO" in concepto
        _data_row(ws1, j, [concepto, monto, base, norma],
                  bold=is_total,
                  font_color=C_AMBER if is_total else C_GRAY_LIGHT,
                  bg_color="0F172A" if j % 2 == 0 else C_DARK_BG)
        ws1.cell(j, 2).number_format = '#,##0.00'

    nota_row = row + 3 + len(imp_rows) + 1
    ws1.merge_cells(f"A{nota_row}:D{nota_row}")
    cn = ws1.cell(nota_row, 1,
        value="⚠ Nota: El ITBIS depende del tipo de actividad económica y no se calcula automáticamente.")
    cn.fill = _fill(C_SURFACE)
    cn.font = _font(color=C_AMBER, size=9, italic=True)
    cn.alignment = _left()

    # ── Hoja 2: Top Movimientos ───────────────────────────
    ws2 = wb.create_sheet("Top Movimientos")
    _set_tab_color(ws2, C_GREEN)
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:C1")
    ws2["A1"].value     = "🌴  OASIS — Top Movimientos"
    ws2["A1"].fill      = _fill(C_DARK_BG)
    ws2["A1"].font      = _font(bold=True, color=C_TEAL, size=14)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 36

    # Top 5 Ingresos
    _title_row(ws2, 2, "🔝  TOP 5 INGRESOS", 3, fg=C_GREEN)
    _header_row(ws2, 3, [("DESCRIPCIÓN", 40), ("MONTO RD$", 20), ("#", 6)], font_color=C_GREEN)
    top_ing = data.get("top_ingresos", [])
    for k, item in enumerate(top_ing, 4):
        desc  = item.get("Descripcion") or item.get("descripcion", "—")
        monto = item.get("Monto_RD")    or item.get("monto_rd", 0)
        _data_row(ws2, k, [desc, monto, k - 3],
                  bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
        ws2.cell(k, 2).font         = _font(bold=True, color=C_GREEN, size=11)
        ws2.cell(k, 2).number_format = '#,##0.00'

    # Top 5 Gastos
    sep = len(top_ing) + 6
    _title_row(ws2, sep, "📉  TOP 5 GASTOS", 3, fg=C_RED)
    _header_row(ws2, sep + 1, [("DESCRIPCIÓN", 40), ("MONTO RD$", 20), ("#", 6)], font_color=C_RED)
    top_gas = data.get("top_gastos", [])
    for k, item in enumerate(top_gas, sep + 2):
        desc  = item.get("Descripcion") or item.get("descripcion", "—")
        monto = item.get("Monto_RD")    or item.get("monto_rd", 0)
        _data_row(ws2, k, [desc, monto, k - sep - 1],
                  bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
        ws2.cell(k, 2).font         = _font(bold=True, color=C_RED, size=11)
        ws2.cell(k, 2).number_format = '#,##0.00'

    # ── Hoja 3: Métodos de Pago ───────────────────────────
    metodos = data.get("metodos_pago", {})
    if metodos:
        ws3 = wb.create_sheet("Métodos de Pago")
        _set_tab_color(ws3, C_BLUE)
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:C1")
        ws3["A1"].value     = "🌴  OASIS — Métodos de Pago"
        ws3["A1"].fill      = _fill(C_DARK_BG)
        ws3["A1"].font      = _font(bold=True, color=C_TEAL, size=14)
        ws3["A1"].alignment = _center()
        ws3.row_dimensions[1].height = 36

        _title_row(ws3, 2, "💳  DISTRIBUCIÓN POR MÉTODO", 3, fg=C_BLUE)
        _header_row(ws3, 3, [
            ("MÉTODO DE PAGO", 30), ("CANTIDAD", 15), ("% DEL TOTAL", 15)
        ], font_color=C_BLUE)

        total_trans = sum(metodos.values())
        for k, (metodo, cantidad) in enumerate(
            sorted(metodos.items(), key=lambda x: x[1], reverse=True), 4
        ):
            pct = (cantidad / total_trans * 100) if total_trans else 0
            _data_row(ws3, k, [metodo, cantidad, f"{pct:.1f}%"],
                      bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
            ws3.cell(k, 2).font = _font(bold=True, color=C_BLUE, size=11)

    # ── Serializar a bytes ────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ══════════════════════════════════════════════════════════
# EXPORTAR ANÁLISIS SANITARIO
# ══════════════════════════════════════════════════════════
def export_sanitario_xlsx(data: dict, nombre_archivo: str = "análisis") -> bytes:
    """
    Genera un .xlsx con los resultados del análisis sanitario.
    """
    wb = Workbook()

    # ── Hoja 1: Resumen General ───────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    _set_tab_color(ws1, C_TEAL)
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value     = "🌴  OASIS — Gestión Sanitaria"
    c.fill      = _fill(C_DARK_BG)
    c.font      = _font(bold=True, color=C_TEAL, size=16)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:D2")
    c2 = ws1["A2"]
    c2.value     = f"Archivo: {nombre_archivo}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c2.fill      = _fill(C_SURFACE)
    c2.font      = _font(color=C_GRAY_MED, size=10, italic=True)
    c2.alignment = _center()

    ws1.row_dimensions[3].height = 8
    _title_row(ws1, 4, "📋  TOTALES GENERALES", 4)

    _header_row(ws1, 5, [
        ("INDICADOR", 30), ("VALOR", 16), ("UNIDAD", 20), ("DESCRIPCIÓN", 36)
    ])

    totales = [
        ("Total Eventos",         data.get("total_eventos", 0),        "eventos",    "Todos los eventos registrados"),
        ("Accidentes Laborales",  data.get("total_accidentes", 0),     "accidentes", "Eventos tipo Accidente"),
        ("Enfermedades Ocup.",    data.get("total_enfermedades", 0),   "casos",      "Eventos tipo Enfermedad"),
        ("Actividades Prevención",data.get("total_prevencion", 0),     "actividades","Eventos tipo Prevención"),
        ("Total Días Perdidos",   data.get("total_dias_perdidos", 0),  "días",       "Días laborales perdidos"),
        ("Costo Total",           data.get("total_costo", 0),          "RD$",        "Suma de todos los costos"),
        ("Costo Prom./Evento",    data.get("costo_promedio_evento", 0),"RD$",        "Costo total / total eventos"),
        ("Casos Resueltos",       data.get("casos_resueltos", 0),      "casos",      "Eventos marcados como resueltos"),
        ("Casos Activos",         data.get("casos_activos", 0),        "casos",      "Eventos aún abiertos"),
        ("Reportados a la ARL",   data.get("reportados_arl", 0),       "casos",      "Obligatorio Ley 87-01"),
    ]

    color_map = {
        "Accidentes Laborales": C_RED,
        "Total Días Perdidos":  C_AMBER,
        "Costo Total":          C_RED,
        "Casos Resueltos":      C_GREEN,
    }

    for i, (ind, val, unit, desc) in enumerate(totales, 6):
        color_val = color_map.get(ind, C_TEAL)
        _data_row(ws1, i, [ind, val, unit, desc],
                  bg_color="0F172A" if i % 2 == 0 else C_DARK_BG)
        ws1.cell(i, 2).font      = _font(bold=True, color=color_val, size=11)
        ws1.cell(i, 2).number_format = '#,##0.00'

    # Indicadores de seguridad
    row_ind = len(totales) + 8
    _title_row(ws1, row_ind, "📐  INDICADORES DE SEGURIDAD (Decreto 522-06)", 4, fg=C_PURPLE)
    _header_row(ws1, row_ind + 1, [
        ("INDICADOR", 30), ("VALOR", 16), ("FÓRMULA", 28), ("BASE NORMATIVA", 28)
    ], font_color=C_PURPLE)

    indicadores = [
        ("Tasa de Frecuencia",  data.get("tasa_frecuencia", 0),
         "Accidentes / h.trab. × 1,000,000", "Decreto 522-06 Art. 4"),
        ("Tasa de Severidad",   data.get("tasa_severidad", 0),
         "Días perdidos / h.trab. × 1,000",  "Decreto 522-06 Art. 4"),
        ("Tasa de Incidencia",  data.get("tasa_incidencia", 0),
         "Eventos / empleados × 100",         "Ley 87-01"),
        ("Índice de Gravedad",  data.get("indice_gravedad", 0),
         "Días perdidos / accidentes",         "Decreto 522-06"),
        ("Total Empleados",     data.get("total_empleados", 0),
         "—",                                  "Base del cálculo"),
        ("Horas Trabajadas",    data.get("horas_trabajadas", 0),
         "Empleados × 2,200 h/año",            "Art. 147 Código Trabajo RD"),
    ]

    for j, (ind, val, formula, norma) in enumerate(indicadores, row_ind + 2):
        _data_row(ws1, j, [ind, val, formula, norma],
                  bg_color="0F172A" if j % 2 == 0 else C_DARK_BG)
        ws1.cell(j, 2).font         = _font(bold=True, color=C_PURPLE, size=11)
        ws1.cell(j, 2).number_format = '#,##0.00'

    # ── Hoja 2: Por Departamento ──────────────────────────
    por_depto = data.get("por_departamento", {})
    if por_depto:
        ws2 = wb.create_sheet("Por Departamento")
        _set_tab_color(ws2, C_BLUE)
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:F1")
        ws2["A1"].value     = "🌴  OASIS — Eventos por Departamento"
        ws2["A1"].fill      = _fill(C_DARK_BG)
        ws2["A1"].font      = _font(bold=True, color=C_TEAL, size=14)
        ws2["A1"].alignment = _center()
        ws2.row_dimensions[1].height = 36

        _title_row(ws2, 2, "🏢  DESGLOSE POR DEPARTAMENTO", 6, fg=C_BLUE)
        _header_row(ws2, 3, [
            ("DEPARTAMENTO", 22), ("EVENTOS", 12), ("ACCIDENTES", 14),
            ("DÍAS PERDIDOS", 15), ("COSTO RD$", 18), ("EMPLEADOS", 12)
        ], font_color=C_BLUE)

        sorted_deptos = sorted(por_depto.items(), key=lambda x: x[1].get("eventos", 0), reverse=True)
        for k, (depto, vals) in enumerate(sorted_deptos, 4):
            _data_row(ws2, k, [
                depto,
                vals.get("eventos", 0),
                vals.get("accidentes", 0),
                vals.get("dias", 0),
                vals.get("costo", 0),
                vals.get("empleados", 0),
            ], bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
            ws2.cell(k, 5).number_format = '#,##0.00'

    # ── Hoja 3: Top Diagnósticos ──────────────────────────
    top_diag = data.get("top_diagnosticos", [])
    if top_diag:
        ws3 = wb.create_sheet("Top Diagnósticos")
        _set_tab_color(ws3, C_AMBER)
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:D1")
        ws3["A1"].value     = "🌴  OASIS — Top Diagnósticos"
        ws3["A1"].fill      = _fill(C_DARK_BG)
        ws3["A1"].font      = _font(bold=True, color=C_TEAL, size=14)
        ws3["A1"].alignment = _center()
        ws3.row_dimensions[1].height = 36

        _title_row(ws3, 2, "🔬  DIAGNÓSTICOS MÁS FRECUENTES", 4, fg=C_AMBER)
        _header_row(ws3, 3, [
            ("DIAGNÓSTICO", 40), ("CASOS", 12), ("DÍAS PERDIDOS", 15), ("COSTO RD$", 18)
        ], font_color=C_AMBER)

        for k, item in enumerate(top_diag, 4):
            _data_row(ws3, k, [
                item.get("diagnostico", "—"),
                item.get("casos", 0),
                item.get("dias", 0),
                item.get("costo", 0),
            ], bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
            ws3.cell(k, 2).font         = _font(bold=True, color=C_AMBER, size=11)
            ws3.cell(k, 4).number_format = '#,##0.00'

    # ── Hoja 4: Tendencia Mensual ─────────────────────────
    tendencia = data.get("tendencia_mensual", [])
    if tendencia:
        ws4 = wb.create_sheet("Tendencia Mensual")
        _set_tab_color(ws4, C_GREEN)
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:F1")
        ws4["A1"].value     = "🌴  OASIS — Tendencia Mensual"
        ws4["A1"].fill      = _fill(C_DARK_BG)
        ws4["A1"].font      = _font(bold=True, color=C_TEAL, size=14)
        ws4["A1"].alignment = _center()
        ws4.row_dimensions[1].height = 36

        _title_row(ws4, 2, "📈  EVOLUCIÓN MES A MES", 6, fg=C_GREEN)
        _header_row(ws4, 3, [
            ("MES", 16), ("TOTAL EVENTOS", 15), ("ACCIDENTES", 14),
            ("ENFERMEDADES", 15), ("DÍAS PERDIDOS", 15), ("COSTO RD$", 18)
        ], font_color=C_GREEN)

        for k, mes in enumerate(tendencia, 4):
            _data_row(ws4, k, [
                mes.get("mes", "—"),
                mes.get("eventos", 0),
                mes.get("accidentes", 0),
                mes.get("enfermedades", 0),
                mes.get("dias", 0),
                mes.get("costo", 0),
            ], bg_color="0F172A" if k % 2 == 0 else C_DARK_BG)
            ws4.cell(k, 6).number_format = '#,##0.00'

    # ── Serializar a bytes ────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()